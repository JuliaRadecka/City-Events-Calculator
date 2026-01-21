# -*- coding: utf-8 -*-
import io
import os
import re
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    # openpyxl provides this class when a cell contains an array formula
    from openpyxl.worksheet.formula import ArrayFormula  # type: ignore
except Exception:  # pragma: no cover
    ArrayFormula = None  # type: ignore


# ============================================================
# App config
# ============================================================

APP_TITLE = "📊 Калькулятор оценки мероприятий"

# We'll auto-detect workbook in working directory, but prefer these names
PREFERRED_BOOK_NAMES = [
    "Калькулятор_оценки_мероприятий_по_городам_все_города.xlsm",
    "Калькулятор_оценки_мероприятий_МТС_Live_по_городам_140825_все_города.xlsm",
    "Calculator.xlsx",
    "Calculator.xlsm",
]

# Excel working blocks (TEMPLATE)
PARAM_R1, PARAM_C1, PARAM_R2, PARAM_C2 = 20, 3, 26, 5     # C20:E26
MAIN_R1, MAIN_C1, MAIN_R2, MAIN_C2 = 27, 3, 134, 15       # C27:O134
BOT_R1, BOT_C1, BOT_R2, BOT_C2 = 135, 3, 146, 18          # C135:R146 (R=18)


# Known fills (ARGB) from the template (fallback)
KNOWN_GREYS = {"FFD9D9D9", "FFBFBFBF", "FFE7E6E6", "FFDDDDDD"}
KNOWN_BLUES = {"FFBDD7EE", "FF9DC3E6", "FFB4C6E7", "FFB7DEE8"}


# ============================================================
# Helpers: Excel basics
# ============================================================


def a1(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def cell_fill_argb(cell) -> str:
    try:
        f = cell.fill
        if f is None or f.patternType is None:
            return ""
        c = f.fgColor
        if c is None:
            return ""
        if c.type == "rgb" and c.rgb:
            return c.rgb
    except Exception:
        return ""
    return ""


def is_grey_input(cell) -> bool:
    argb = cell_fill_argb(cell)
    if argb in KNOWN_GREYS:
        return True
    # fallback: non-white non-blue treated as grey input in this template
    if argb and argb not in KNOWN_BLUES and argb != "FFFFFFFF":
        return True
    return False


def is_blue_auto(cell) -> bool:
    return cell_fill_argb(cell) in KNOWN_BLUES


def coerce_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)) and not isinstance(x, bool):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def safe_excel_scalar(v: Any) -> Any:
    """Convert openpyxl-specific value types to python scalars for Streamlit.

    Critical: openpyxl ArrayFormula objects crash Streamlit data_editor/pyarrow.
    We never pass them to the UI.
    """
    if v is None:
        return None

    if ArrayFormula is not None and isinstance(v, ArrayFormula):
        # keep the formula text if needed; for UI we generally show computed value, not formula
        try:
            return f"={v.text}" if getattr(v, "text", None) else ""
        except Exception:
            return ""

    # Some openpyxl objects stringify weirdly; enforce primitives where possible
    if isinstance(v, (int, float, str, bool)):
        return v

    try:
        return str(v)
    except Exception:
        return ""


# ============================================================
# Workbook discovery
# ============================================================


def find_workbook_path() -> Path:
    here = Path(__file__).resolve().parent
    for name in PREFERRED_BOOK_NAMES:
        p = here / name
        if p.exists():
            return p
    # fall back to first xlsx/xlsm in folder
    candidates = list(here.glob("*.xlsm")) + list(here.glob("*.xlsx"))
    if candidates:
        return candidates[0]
    # try CWD
    cwd = Path.cwd()
    for name in PREFERRED_BOOK_NAMES:
        p = cwd / name
        if p.exists():
            return p
    candidates = list(cwd.glob("*.xlsm")) + list(cwd.glob("*.xlsx"))
    if candidates:
        return candidates[0]
    raise FileNotFoundError("Не найден Excel-файл (.xlsm/.xlsx) рядом с приложением.")


# ============================================================
# Reference lists ("Списки")
# ============================================================


def extract_city_list(ws_lists) -> List[str]:
    cities: List[str] = []
    r = 2
    while True:
        v = ws_lists[f"A{r}"].value
        if v is None or str(v).strip() == "":
            break
        cities.append(str(v).strip())
        r += 1
    # protect from accidental blanks
    return [c for c in cities if c.strip()]


def extract_type_ploshadki(ws_lists) -> List[str]:
    opts: List[str] = []
    for r in range(2, 200):
        v = ws_lists[f"W{r}"].value
        if v is None or str(v).strip() == "":
            break
        opts.append(str(v).strip())
    return [o for o in opts if o.strip()]


def extract_union_formats(ws_lists) -> List[str]:
    # union of all format columns used in the template lists
    cols = ["B", "C", "D", "E", "F", "H", "I", "K", "L", "M", "N", "O", "P", "Q"]
    out: set[str] = set()
    for col in cols:
        empty_streak = 0
        for r in range(2, 500):
            v = ws_lists[f"{col}{r}"].value
            if v is None or str(v).strip() == "":
                empty_streak += 1
                if r > 30 and empty_streak > 25:
                    break
                continue
            empty_streak = 0
            out.add(str(v).strip())
    return sorted(out)


def extract_simple_col(ws, col_letter: str, start_row: int = 2, max_rows: int = 500) -> List[str]:
    out: List[str] = []
    for r in range(start_row, max_rows + 1):
        v = ws[f"{col_letter}{r}"].value
        if v is None or str(v).strip() == "":
            continue
        out.append(str(v).strip())
    return out


# ============================================================
# Excel-like formula evaluation engine (minimal)
# (kept from draft: enough for E21 VLOOKUP + some simple formulas)
# ============================================================


@dataclass
class EvalContext:
    wb: Any
    sheet: str
    values: Dict[str, Any]  # A1 -> scalar


_CELL_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


def _strip_dollars(ref: str) -> str:
    return ref.replace("$", "")


def _parse_range(rng: str) -> Tuple[str, str]:
    a, b = rng.split(":")
    return _strip_dollars(a), _strip_dollars(b)


def _expand_range(a_ref: str, b_ref: str) -> List[str]:
    col_a = re.findall(r"[A-Z]{1,3}", a_ref)[0]
    row_a = int(re.findall(r"\d+", a_ref)[0])
    col_b = re.findall(r"[A-Z]{1,3}", b_ref)[0]
    row_b = int(re.findall(r"\d+", b_ref)[0])
    ca = column_index_from_string(col_a)
    cb = column_index_from_string(col_b)

    out: List[str] = []
    for r in range(min(row_a, row_b), max(row_a, row_b) + 1):
        for c in range(min(ca, cb), max(ca, cb) + 1):
            out.append(a1(r, c))
    return out


def _get(ctx: EvalContext, ref: str) -> Any:
    ref = _strip_dollars(ref)
    return ctx.values.get(ref)


def _to_number(v: Any) -> Optional[float]:
    return coerce_float(v)


def _excel_sum(args: List[Any]) -> float:
    total = 0.0
    for v in args:
        if isinstance(v, list):
            total += _excel_sum(v)
        else:
            num = _to_number(v)
            if num is not None:
                total += num
    return total


def _excel_if(cond: Any, a_val: Any, b_val: Any) -> Any:
    return a_val if bool(cond) else b_val


def _excel_vlookup(ctx: EvalContext, lookup_value: Any, table_range: str, col_index: int) -> Any:
    sheet = ctx.sheet
    rng = table_range
    if "!" in table_range:
        sh, rng = table_range.split("!", 1)
        sh = sh.strip().strip("'")
        sh = sh.replace("[1]", "")
        sheet = sh

    a_ref, b_ref = _parse_range(rng)
    ws = ctx.wb[sheet]

    a_col = re.findall(r"[A-Z]{1,3}", a_ref)[0]
    a_row = int(re.findall(r"\d+", a_ref)[0])
    b_col = re.findall(r"[A-Z]{1,3}", b_ref)[0]
    b_row = int(re.findall(r"\d+", b_ref)[0])
    ca = column_index_from_string(a_col)

    lookup_str = str(lookup_value).strip() if lookup_value is not None else ""

    for r in range(a_row, b_row + 1):
        key = ws.cell(r, ca).value
        if str(key).strip() == lookup_str:
            target_c = ca + (col_index - 1)
            return ws.cell(r, target_c).value
    return None


def _transform_formula(formula: str) -> str:
    f = formula.strip()
    if f.startswith("="):
        f = f[1:]

    f = f.replace("^", "**")
    f = re.sub(r"\bTRUE\b", "True", f, flags=re.IGNORECASE)
    f = re.sub(r"\bFALSE\b", "False", f, flags=re.IGNORECASE)
    f = f.replace("[1]", "")

    # ranges A1:B2 -> RNG('A1','B2')
    f = re.sub(
        r"(\$?[A-Z]{1,3}\$?\d+)\s*:\s*(\$?[A-Z]{1,3}\$?\d+)",
        lambda m: f"RNG('{_strip_dollars(m.group(1))}','{_strip_dollars(m.group(2))}')",
        f,
    )

    # protect quoted strings
    strings: List[str] = []

    def protect_str(m):
        strings.append(m.group(0))
        return f"__STR{len(strings)-1}__"

    f = re.sub(r'"[^"]*"', protect_str, f)

    def repl_cell(m):
        ref = _strip_dollars(m.group(0))
        return f"CELL('{ref}')"

    f = _CELL_RE.sub(repl_cell, f)

    for i, s in enumerate(strings):
        f = f.replace(f"__STR{i}__", s)

    f = re.sub(r"\bSUM\s*\(", "SUM(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bIF\s*\(", "IF(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bVLOOKUP\s*\(", "VLOOKUP(", f, flags=re.IGNORECASE)

    return f


def eval_formula(ctx: EvalContext, formula: str) -> Any:
    expr = _transform_formula(formula)

    def CELL(ref: str) -> Any:
        return _get(ctx, ref)

    def RNG(a_ref: str, b_ref: str) -> List[Any]:
        return [_get(ctx, r) for r in _expand_range(a_ref, b_ref)]

    def SUM(*args):
        return _excel_sum(list(args))

    def IF(cond, a_val, b_val):
        return _excel_if(cond, a_val, b_val)

    def VLOOKUP(lookup_value, table_range, col_index, exact=True):
        return _excel_vlookup(ctx, lookup_value, table_range, int(col_index))

    env = {"CELL": CELL, "RNG": RNG, "SUM": SUM, "IF": IF, "VLOOKUP": VLOOKUP}

    try:
        return eval(expr, {"__builtins__": {}}, env)
    except Exception:
        # Many template formulas (SUMPRODUCT etc.) are not supported in this minimal engine.
        return None


def recalc_block(ctx: EvalContext, ws, addr_list: List[str], max_iters: int = 12) -> None:
    formulas: Dict[str, str] = {}
    for addr in addr_list:
        v = ws[addr].value
        if isinstance(v, str) and v.startswith("="):
            formulas[addr] = v
        elif ArrayFormula is not None and isinstance(v, ArrayFormula):
            # openpyxl array formula
            txt = getattr(v, "text", "")
            formulas[addr] = f"={txt}" if txt else ""

    for _ in range(max_iters):
        changed = 0
        for addr, f in formulas.items():
            if not f:
                continue
            new_val = eval_formula(ctx, f)
            old_val = ctx.values.get(addr)
            if new_val != old_val:
                ctx.values[addr] = new_val
                changed += 1
        if changed == 0:
            break


# ============================================================
# Template reading
# ============================================================


def read_block_style(ws_style, r1: int, c1: int, r2: int, c2: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    grey: List[List[bool]] = []
    blue: List[List[bool]] = []
    for r in range(r1, r2 + 1):
        rg: List[bool] = []
        rb: List[bool] = []
        for c in range(c1, c2 + 1):
            cell = ws_style.cell(r, c)
            rg.append(is_grey_input(cell))
            rb.append(is_blue_auto(cell))
        grey.append(rg)
        blue.append(rb)
    cols = [get_column_letter(c) for c in range(c1, c2 + 1)]
    idx = list(range(r1, r2 + 1))
    return pd.DataFrame(grey, columns=cols, index=idx), pd.DataFrame(blue, columns=cols, index=idx)


def style_df(df: pd.DataFrame, grey_mask: pd.DataFrame, blue_mask: pd.DataFrame):
    def _apply(_):
        out = pd.DataFrame("", index=df.index, columns=df.columns)
        out[grey_mask] = "background-color: #D9D9D9;"  # light grey
        out[blue_mask] = "background-color: #9DC3E6;"  # Excel-like light blue
        return out

    return df.style.apply(_apply, axis=None)


def inject_css():
    st.markdown(
        """
<style>
  div[data-testid="stVerticalBlock"] div[data-testid="stTextInput"] input,
  div[data-testid="stVerticalBlock"] div[data-testid="stNumberInput"] input,
  div[data-testid="stVerticalBlock"] div[data-testid="stSelectbox"] div {
    max-width: 260px;
  }
  .block-title { font-weight: 700; font-size: 1.05rem; margin-bottom: 0.35rem; }
  .calc-area { padding: 0.5rem 0.75rem; border: 1px solid rgba(49, 51, 63, 0.2); border-radius: 0.5rem; }

  /* Auto-calculated (read-only) fields: solid light-blue fill.
     Streamlit draws the visible background on the wrapper div, so we color both
     the wrapper and the input to avoid the "half-filled" effect. */
  div[data-testid="stTextInput"] div[data-baseweb="input"]:has(input:disabled),
  div[data-testid="stNumberInput"] div[data-baseweb="input"]:has(input:disabled) {
    background: rgba(60, 140, 220, 0.35) !important;
    border-radius: 6px !important;
  }
  div[data-testid="stTextInput"] input:disabled,
  div[data-testid="stNumberInput"] input:disabled {
    -webkit-text-fill-color: rgba(255, 255, 255, 0.95) !important;
    color: rgba(255, 255, 255, 0.95) !important;
    background: rgba(60, 140, 220, 0.35) !important;
  }
</style>
""",
        unsafe_allow_html=True,
    )


@st.cache_resource(show_spinner=False)
def load_wb():
    path = find_workbook_path()
    wb = load_workbook(path, data_only=False, keep_vba=True)
    if "TEMPLATE" in wb.sheetnames:
        ws = wb["TEMPLATE"]
    else:
        # fallback: first sheet
        ws = wb[wb.sheetnames[0]]
    if "Списки" not in wb.sheetnames:
        raise RuntimeError("Не найден лист 'Списки' в книге.")
    return path, wb, ws, wb["Списки"]


def init_calc_values(ws) -> Dict[str, Any]:
    vals: Dict[str, Any] = {}
    ranges = [
        (PARAM_R1, PARAM_C1, PARAM_R2, PARAM_C2),
        (MAIN_R1, MAIN_C1, MAIN_R2, MAIN_C2),
        (BOT_R1, BOT_C1, BOT_R2, BOT_C2),
        (21, 5, 21, 5),  # E21
    ]
    seen: set[str] = set()
    for r1, c1, r2, c2 in ranges:
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                addr = a1(r, c)
                if addr in seen:
                    continue
                seen.add(addr)
                v = ws[addr].value
                # if formula: store placeholder None; computed will be injected after recalc
                if isinstance(v, str) and v.startswith("="):
                    vals[addr] = None
                elif ArrayFormula is not None and isinstance(v, ArrayFormula):
                    vals[addr] = None
                else:
                    vals[addr] = safe_excel_scalar(v)
    return vals


def build_media_display(ws, vals_map: Dict[str, Any]) -> Tuple[pd.DataFrame, List[str], pd.DataFrame, pd.DataFrame, List[List[str]]]:
    """Build display dataframe for media table (rows 28..134, cols C..O).

    Returns: display_df, headers, grey_mask, blue_mask, addr_matrix
    """
    # Headers from row 27 (C..O)
    headers = []
    for c in range(MAIN_C1, MAIN_C2 + 1):
        v = ws.cell(MAIN_R1, c).value
        h = str(v).strip() if v is not None and str(v).strip() else get_column_letter(c)
        headers.append(h)

    grey_mask, blue_mask = read_block_style(ws, MAIN_R1 + 1, MAIN_C1, MAIN_R2, MAIN_C2)
    grey_mask.columns = headers
    blue_mask.columns = headers

    rows: List[List[Any]] = []
    addr_matrix: List[List[str]] = []

    for r in range(MAIN_R1 + 1, MAIN_R2 + 1):
        row_vals: List[Any] = []
        row_addrs: List[str] = []
        i = r - (MAIN_R1 + 1)
        for j, c in enumerate(range(MAIN_C1, MAIN_C2 + 1)):
            addr = a1(r, c)
            row_addrs.append(addr)

            # If cell is formula in template, show computed value (from vals_map), else raw input
            template_v = ws[addr].value
            if (isinstance(template_v, str) and template_v.startswith("=")) or (
                ArrayFormula is not None and isinstance(template_v, ArrayFormula)
            ):
                v = vals_map.get(addr)
            else:
                v = vals_map.get(addr)

            # Normalize for UI
            if v is None:
                v = ""
            row_vals.append(v)

        rows.append(row_vals)
        addr_matrix.append(row_addrs)

    df = pd.DataFrame(rows, columns=headers)
    return df, headers, grey_mask, blue_mask, addr_matrix


def split_media_sections(df: pd.DataFrame, headers: List[str]) -> List[Tuple[str, pd.DataFrame]]:
    """Split media table into activation blocks by rows where first column starts with 'Тип активации'.

    In Excel these are inside column C of the C..O block.
    """
    if not headers:
        return [("Медиа факторы", df)]
    first_col = headers[0]

    # row indices in the df (0-based) where activation header occurs
    act_rows = []
    for idx, val in enumerate(df[first_col].astype(str).tolist()):
        if str(val).strip().lower().startswith("тип активации"):
            act_rows.append(idx)

    if not act_rows:
        return [("Медиа факторы", df)]

    blocks: List[Tuple[str, pd.DataFrame]] = []
    for k, start in enumerate(act_rows):
        end = act_rows[k + 1] if k + 1 < len(act_rows) else len(df)
        title = str(df.iloc[start][first_col]).strip()
        block_df = df.iloc[start:end].copy()
        blocks.append((title, block_df))
    return blocks


# ============================================================
# Session state: multi-geo
# ============================================================


def ensure_state():
    if "geo_tabs" not in st.session_state:
        st.session_state["geo_tabs"] = []  # list of {id, geo}
    if "active_geo_add" not in st.session_state:
        st.session_state["active_geo_add"] = None


def add_geo_tab(geo_name: str):
    geo_name = str(geo_name).strip()
    if not geo_name:
        return
    tab_id = str(uuid.uuid4())[:8]
    st.session_state["geo_tabs"].append({"id": tab_id, "geo": geo_name})

    # per-tab calc values + media editor storage
    st.session_state[f"{tab_id}__calc_values"] = None  # init later
    st.session_state[f"{tab_id}__media_df"] = None
    st.session_state[f"{tab_id}__params"] = {
        "geo": geo_name,
        "type_pl": None,
        "days": 0.0,
        "period": 0.0,
        "visitors": 0.0,
        "ticket_price": 0.0,
        "gmv": 0.0,
        "fee": 0.0,
        "widget_tickets": 0.0,
        "integ_fee": 0.0,
        "production": 0.0,
    }


# ============================================================
# UI rendering
# ============================================================


def render_top_blocks(tab_id: str, cities: List[str], type_pl: List[str], ca_value: Any):
    params = st.session_state[f"{tab_id}__params"]

    left, mid1, mid2, right = st.columns([1, 1, 1, 1], gap="large")

    # Parameters
    with left:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">⚙️Параметры</div>', unsafe_allow_html=True)

        st.text_input(
            "ЦА (унифицированная аудитория для всех медиа, тыс. 16+)",
            value="" if ca_value in (None, "") else str(ca_value),
            disabled=True,
            key=f"{tab_id}__ca",
        )

        # Geo
        if params.get("geo") in cities:
            idx = cities.index(params.get("geo"))
        else:
            idx = 0
        params["geo"] = st.selectbox("ГЕО", options=cities, index=idx, key=f"{tab_id}__geo")

        p1, p2 = st.columns([1, 1])
        with p1:
            # Type площадки
            if params.get("type_pl") in type_pl:
                idx2 = type_pl.index(params.get("type_pl"))
            else:
                idx2 = 0
            params["type_pl"] = st.selectbox(
                "Тип площадки",
                options=type_pl,
                index=idx2,
                key=f"{tab_id}__type_pl",
            )
        with p2:
            params["days"] = st.number_input(
                "Кол-во дней на фестивале/площадке",
                value=float(params.get("days", 0.0) or 0.0),
                step=1.0,
                key=f"{tab_id}__days",
            )

        params["period"] = st.number_input(
            "Общий период размещения",
            value=float(params.get("period", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__period",
        )
        params["visitors"] = st.number_input(
            "План посетителей (в тыс. человек)",
            value=float(params.get("visitors", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__visitors",
        )

        st.markdown('</div>', unsafe_allow_html=True)

    # Planned result
    with mid1:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">ПЛАНОВЫЙ РЕЗУЛЬТАТ МЕРОПРИЯТИЯ</div>', unsafe_allow_html=True)
        st.text_input(
            "Кол-во посетителей всего, тыс.",
            value=str(params.get("visitors", 0.0) or 0.0),
            disabled=True,
            key=f"{tab_id}__visitors_ro",
        )
        params["ticket_price"] = st.number_input(
            "Средняя цена билета",
            value=float(params.get("ticket_price", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__ticket_price",
        )
        params["gmv"] = st.number_input(
            "GMV",
            value=float(params.get("gmv", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__gmv",
        )
        params["fee"] = st.number_input(
            "Агентская комиссия",
            value=float(params.get("fee", 0.0) or 0.0),
            step=0.1,
            key=f"{tab_id}__fee",
        )
        params["widget_tickets"] = st.number_input(
            "Количество проданных билетов через виджит/витрину",
            value=float(params.get("widget_tickets", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__widget_tickets",
        )
        st.markdown('</div>', unsafe_allow_html=True)

    # Budget
    with mid2:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">💰 БЮДЖЕТ</div>', unsafe_allow_html=True)
        params["integ_fee"] = st.number_input(
            "Интеграционный платеж (организатору)",
            value=float(params.get("integ_fee", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__integ_fee",
        )
        params["production"] = st.number_input(
            "Продакшен фото-зоны и лайтбокса, букинг секретных артистов",
            value=float(params.get("production", 0.0) or 0.0),
            step=1.0,
            key=f"{tab_id}__production",
        )
        total_budget = float(params.get("integ_fee", 0.0) or 0.0) + float(params.get("production", 0.0) or 0.0)
        params["total_budget"] = total_budget
        st.text_input("Общий бюджет", value=f"{total_budget}", disabled=True, key=f"{tab_id}__total_budget")
        st.markdown('</div>', unsafe_allow_html=True)

    # Efficiency
    with right:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">📈ЭФФЕКТИВНОСТЬ</div>', unsafe_allow_html=True)
        eff = st.session_state.get(f"{tab_id}__eff", {})
        st.text_input("Стоимость привлеченного клиента", value=str(eff.get("cpa", "")), disabled=True, key=f"{tab_id}__cpa")
        st.text_input("Стоимость за контакт", value=str(eff.get("cpc", "")), disabled=True, key=f"{tab_id}__cpc")
        st.text_input("Стоимость за охваченного пользователя", value=str(eff.get("cpr", "")), disabled=True, key=f"{tab_id}__cpr")
        st.text_input("Стоимость за посетителя мероприятия", value=str(eff.get("cpv", "")), disabled=True, key=f"{tab_id}__cpv")
        st.markdown('</div>', unsafe_allow_html=True)


def render_media_blocks(tab_id: str, ws, ws_lists, union_formats: List[str]):
    # init calc values for this tab if needed
    if st.session_state[f"{tab_id}__calc_values"] is None:
        st.session_state[f"{tab_id}__calc_values"] = init_calc_values(ws)

    vals_map: Dict[str, Any] = st.session_state[f"{tab_id}__calc_values"]

    df, headers, grey_mask, blue_mask, addr_matrix = build_media_display(ws, vals_map)

    # If we already have edited df (stored) -> start from it, but keep computed cells from vals_map
    stored = st.session_state.get(f"{tab_id}__media_df")
    if isinstance(stored, pd.DataFrame) and not stored.empty:
        # Align columns; keep same shape
        try:
            df = stored.copy()
        except Exception:
            pass

    # Build blocks by "Тип активации" rows
    blocks = split_media_sections(df, headers)

    # Column config: formats as selectbox
    fmt_cols = [h for h in headers if "ФОРМ" in str(h).upper()]
    column_config = {}
    for h in fmt_cols:
        column_config[h] = st.column_config.SelectboxColumn(h, options=union_formats, required=False)

    # NOTE: st.data_editor cannot disable per-cell; we'll accept edits but later write back only grey cells.
    for title, block_df in blocks:
        st.markdown(f"**{title}**")
        key = f"{tab_id}__media_editor__{abs(hash(title))}"
        edited = st.data_editor(
            block_df.fillna(""),
            use_container_width=True,
            height=min(720, 26 * (len(block_df) + 1)),
            column_config=column_config,
            key=key,
        )

        # Write edited block back into the full df
        df.loc[block_df.index, :] = edited

    # Persist full edited df
    st.session_state[f"{tab_id}__media_df"] = df

    return df, headers, grey_mask, blue_mask, addr_matrix


def apply_user_inputs_and_recalc(tab_id: str, wb, ws):
    """Write parameters + edited grey cells into calc_values, then run minimal recalc."""
    vals: Dict[str, Any] = st.session_state[f"{tab_id}__calc_values"].copy()
    params = st.session_state[f"{tab_id}__params"]

    # Parameters mapping (template): E22:E26
    vals["E22"] = params.get("geo")
    vals["E23"] = params.get("type_pl")
    vals["E24"] = float(params.get("days", 0.0) or 0.0)
    vals["E25"] = float(params.get("period", 0.0) or 0.0)
    vals["E26"] = float(params.get("visitors", 0.0) or 0.0)

    # Media table edits: write back ONLY grey cells
    edited_df = st.session_state.get(f"{tab_id}__media_df")
    if isinstance(edited_df, pd.DataFrame) and not edited_df.empty:
        df, headers, grey_mask, _blue_mask, addr_matrix = build_media_display(ws, vals)
        # Use edited_df but same indexing; addr_matrix corresponds to full df index
        for i in range(len(edited_df)):
            for j in range(len(headers)):
                if not bool(grey_mask.iloc[i, j]):
                    continue
                addr = addr_matrix[i][j]
                v = edited_df.iloc[i, j]
                if v == "":
                    v = None
                vals[addr] = v

    ctx = EvalContext(wb=wb, sheet=ws.title, values=vals)

    # Recalc only a small safe set: E21 (ЦА) depends on E22 and 'ЦА по городам'
    # plus any formulas in PARAM/MEDIA/BOTTOM blocks (will mostly evaluate to None if unsupported)
    calc_cells: List[str] = ["E21"]

    for r in range(PARAM_R1, PARAM_R2 + 1):
        for c in range(PARAM_C1, PARAM_C2 + 1):
            addr = a1(r, c)
            v = ws[addr].value
            if isinstance(v, str) and v.startswith("="):
                calc_cells.append(addr)
            elif ArrayFormula is not None and isinstance(v, ArrayFormula):
                calc_cells.append(addr)

    for r in range(MAIN_R1, MAIN_R2 + 1):
        for c in range(MAIN_C1, MAIN_C2 + 1):
            addr = a1(r, c)
            v = ws[addr].value
            if isinstance(v, str) and v.startswith("="):
                calc_cells.append(addr)
            elif ArrayFormula is not None and isinstance(v, ArrayFormula):
                calc_cells.append(addr)

    for r in range(BOT_R1, BOT_R2 + 1):
        for c in range(BOT_C1, BOT_C2 + 1):
            addr = a1(r, c)
            v = ws[addr].value
            if isinstance(v, str) and v.startswith("="):
                calc_cells.append(addr)
            elif ArrayFormula is not None and isinstance(v, ArrayFormula):
                calc_cells.append(addr)

    recalc_block(ctx, ws, calc_cells)

    st.session_state[f"{tab_id}__calc_values"] = ctx.values

    # Efficiency metric: CPV = total budget / visitors
    visitors = coerce_float(params.get("visitors")) or 0.0
    total_budget = float(params.get("integ_fee", 0.0) or 0.0) + float(params.get("production", 0.0) or 0.0)
    eff = st.session_state.get(f"{tab_id}__eff", {})
    eff["cpv"] = (total_budget / visitors) if visitors else ""
    st.session_state[f"{tab_id}__eff"] = eff


def render_bottom_block(tab_id: str, ws):
    st.subheader("Итоги / нижний блок")
    vals_map: Dict[str, Any] = st.session_state.get(f"{tab_id}__calc_values") or {}

    bot_rows: List[List[Any]] = []
    for r in range(BOT_R1, BOT_R2 + 1):
        row = []
        for c in range(BOT_C1, BOT_C2 + 1):
            addr = a1(r, c)
            v = vals_map.get(addr)
            row.append("" if v is None else v)
        bot_rows.append(row)

    bot_cols = [get_column_letter(c) for c in range(BOT_C1, BOT_C2 + 1)]
    bot_df = pd.DataFrame(bot_rows, columns=bot_cols, index=list(range(BOT_R1, BOT_R2 + 1)))

    bot_grey, bot_blue = read_block_style(ws, BOT_R1, BOT_C1, BOT_R2, BOT_C2)
    st.dataframe(style_df(bot_df, bot_grey, bot_blue), use_container_width=True, height=360)


def build_export_xlsx(path: Path, wb, ws, tab_id: str) -> bytes:
    # Create a copy workbook for export (values only)
    wb_out = load_workbook(path, data_only=False, keep_vba=False)

    # Ensure TEMPLATE is present
    if ws.title in wb_out.sheetnames:
        ws_out = wb_out[ws.title]
    elif "TEMPLATE" in wb_out.sheetnames:
        ws_out = wb_out["TEMPLATE"]
    else:
        ws_out = wb_out[wb_out.sheetnames[0]]

    vals_map: Dict[str, Any] = st.session_state.get(f"{tab_id}__calc_values") or {}

    # Write values into visible blocks (parameters, main, bottom)
    def _write_range(r1, c1, r2, c2):
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                addr = a1(r, c)
                if addr not in ws_out:
                    continue
                if addr in vals_map and vals_map[addr] is not None:
                    ws_out[addr].value = vals_map[addr]

    _write_range(PARAM_R1, PARAM_C1, PARAM_R2, PARAM_C2)
    _write_range(21, 5, 21, 5)
    _write_range(MAIN_R1, MAIN_C1, MAIN_R2, MAIN_C2)
    _write_range(BOT_R1, BOT_C1, BOT_R2, BOT_C2)

    out = io.BytesIO()
    wb_out.save(out)
    out.seek(0)
    return out.getvalue()


# ============================================================
# Main
# ============================================================


def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    inject_css()
    st.title(APP_TITLE)

    ensure_state()

    try:
        book_path, wb, ws_template, ws_lists = load_wb()
    except Exception as e:
        st.error("Ошибка при загрузке Excel-книги.")
        st.exception(e)
        st.stop()

    # Reference lists
    cities = extract_city_list(ws_lists)
    type_pl = extract_type_ploshadki(ws_lists)
    union_formats = extract_union_formats(ws_lists)

    # Fix: no blank option in add-city select
    cities_for_add = [c for c in cities if c.strip()]

    # Initial tab if none
    if not st.session_state["geo_tabs"]:
        if cities_for_add:
            add_geo_tab(cities_for_add[0])
        else:
            add_geo_tab("Москва")

    # Add city UI (top) — no empty row
    top_left, top_mid, top_right = st.columns([3, 1, 6])
    with top_left:
        st.caption("Добавить город")
        st.session_state["active_geo_add"] = st.selectbox(
            "",
            options=cities_for_add,
            index=min(0, len(cities_for_add) - 1) if cities_for_add else 0,
            label_visibility="collapsed",
            key="geo_add_select",
        )
    with top_mid:
        st.write(" ")
        if st.button("Добавить", key="add_geo_btn"):
            add_geo_tab(st.session_state["active_geo_add"])

    # Tabs
    tab_titles = [t["geo"] for t in st.session_state["geo_tabs"]] + ["UFO"]
    tabs = st.tabs(tab_titles)

    # UFO aggregation placeholder
    ufo_rows: List[Dict[str, Any]] = []

    for idx, t in enumerate(st.session_state["geo_tabs"]):
        tab_id = t["id"]
        geo_name = t["geo"]

        with tabs[idx]:
            st.subheader(geo_name)

            # Ensure per-tab values
            if st.session_state[f"{tab_id}__calc_values"] is None:
                st.session_state[f"{tab_id}__calc_values"] = init_calc_values(ws_template)

            vals_map = st.session_state[f"{tab_id}__calc_values"]
            ca_val = vals_map.get("E21")

            # Top blocks like draft
            render_top_blocks(tab_id, cities, type_pl, ca_val)

            st.divider()
            st.subheader("Медиа факторы")

            # Media table blocks
            render_media_blocks(tab_id, ws_template, ws_lists, union_formats)

            # ТВ/Радио selector placeholder (not implemented yet; keeps UI stable)
            with st.expander("ТВ и Радио: выбор каналов / радиостанций", expanded=False):
                if "ТВ Каналы" in wb.sheetnames:
                    tv_channels = extract_simple_col(wb["ТВ Каналы"], "A")
                else:
                    tv_channels = []
                if "Радиостанции" in wb.sheetnames:
                    radio_stations = extract_simple_col(wb["Радиостанции"], "A")
                else:
                    radio_stations = []
                st.multiselect("ТВ: каналы", options=tv_channels, key=f"{tab_id}__tv_channels")
                st.multiselect("Радио: радиостанции", options=radio_stations, key=f"{tab_id}__radio_stations")

            st.divider()
            c1, c2 = st.columns([1, 3])
            with c1:
                if st.button("🧮 Рассчитать", type="primary", key=f"{tab_id}__calc"):
                    apply_user_inputs_and_recalc(tab_id, wb, ws_template)
                    st.success("Расчёт выполнен (частично: поддерживаются базовые формулы, включая ЦА).")

            with c2:
                xlsx_bytes = build_export_xlsx(book_path, wb, ws_template, tab_id)
                st.download_button(
                    label="💾 Скачать файл (.xlsx)",
                    data=xlsx_bytes,
                    file_name=f"Калькулятор_{geo_name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{tab_id}__dl",
                )

            # Нижний блок "Итоги / нижний блок" (как в Excel под кнопками)
            # По ТЗ перенесён наверх, поэтому снизу больше не рендерим.
            # render_bottom_block(tab_id, ws_template)

            # UFO row (minimal for now)
            params = st.session_state[f"{tab_id}__params"]
            total_budget = float(params.get("integ_fee", 0.0) or 0.0) + float(params.get("production", 0.0) or 0.0)
            ufo_rows.append(
                {
                    "ГЕО": params.get("geo"),
                    "Тип площадки": params.get("type_pl"),
                    "Период": params.get("period"),
                    "Дней": params.get("days"),
                    "Посетители, тыс.": params.get("visitors"),
                    "Бюджет": total_budget,
                    "ЦА (E21)": st.session_state.get(f"{tab_id}__calc_values", {}).get("E21"),
                }
            )

    # UFO tab
    with tabs[-1]:
        st.subheader("UFO")
        st.caption("Сводка по вкладкам ГЕО (в Excel это заполняется через 'Итоговые показатели по городам').")
        ufo_df = pd.DataFrame(ufo_rows)
        st.dataframe(ufo_df, use_container_width=True)
        st.download_button(
            label="Скачать UFO (CSV)",
            data=ufo_df.to_csv(index=False).encode("utf-8-sig"),
            file_name="UFO.csv",
            mime="text/csv",
            key="ufo_csv",
        )


if __name__ == "__main__":
    main()
