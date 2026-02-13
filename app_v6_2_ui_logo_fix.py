import os
import base64
import re
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode


# ----------------------------
# Helpers
# ----------------------------
def norm(x) -> str:
    """Normalization for matching only (NOT for UI)."""
    s = "" if x is None else str(x)
    s = s.replace("\n", " ")
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def safe_display_value(v):
    """Excel data_only=True may return None. We show empty for None."""
    return "" if v is None else v


def file_mtime(path: str) -> float:
    try:
        return os.path.getmtime(path)
    except Exception:
        return 0.0


# NOTE: to ensure dropdown lists refresh when Excel changes while Streamlit is running,
# we key cache by the file mtime.
@st.cache_data(show_spinner=False)
def read_single_column_list_cached(wb_path: str, sheet_name: str, col: int, mtime: float) -> List[str]:
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    out = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() != "":
            out.append(str(v).strip())
    return out


@st.cache_data(show_spinner=False)
def read_formats_list_cached(wb_path: str, sheet_name: str, mtime: float) -> List[str]:
    """
    Sheet 'Форматы' can be two columns [Описание, Форматы].
    We take unique values from column 2 in stable order.
    """
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    vals = []
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v is not None and str(v).strip() != "":
            vals.append(str(v).strip())
    seen = set()
    out = []
    for x in vals:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


@st.cache_data(show_spinner=False)
def read_lists_options_cached(wb_path: str, sheet_name: str, mtime: float) -> Dict[str, List[str]]:
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    options: Dict[str, List[str]] = {}
    for r in range(2, ws.max_row + 1):
        f = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if f is None or v is None:
            continue
        f_str = str(f).strip()
        v_str = str(v).strip()
        if f_str and v_str:
            options.setdefault(f_str, []).append(v_str)
    return options


def find_media_headers_and_groups(wb_path: str, sheet_name: str = "Медиа факторы") -> Tuple[List[str], Dict[str, int]]:
    """
    Draft-mode allowed read:
    - headers row1 A..M (we will show B..M)
    - column A values to count rows per activation type
    """
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, 14)]
    headers = ["" if h is None else str(h) for h in headers]

    group_counts: Dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        key = str(a).strip()
        if key:
            group_counts[key] = group_counts.get(key, 0) + 1

    return headers, group_counts


def read_media_factors_truth(wb_path: str, sheet_name: str = "Медиа факторы") -> Tuple[List[str], pd.DataFrame]:
    """Read full A..M (truth) with data_only=True; return headers and dataframe."""
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]

    headers = [ws.cell(1, c).value for c in range(1, 14)]
    headers = ["" if h is None else str(h) for h in headers]

    rows = []
    for r in range(2, ws.max_row + 1):
        row = [safe_display_value(ws.cell(r, c).value) for c in range(1, 14)]
        if all(v == "" for v in row):
            continue
        rows.append(row)

    df = pd.DataFrame(rows, columns=headers)
    return headers, df


def read_filters_rows_truth(wb_path: str, sheet_name: str = "Фильтры") -> List[Dict[str, object]]:
    """
    Read sheet 'Фильтры' as row list (truth).
    Expected headers: A=Блок, B=Название, C=Данные
    """
    wb = load_workbook(wb_path, data_only=True)
    ws = wb[sheet_name]
    out = []
    for r in range(2, ws.max_row + 1):
        block = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        val = ws.cell(r, 3).value
        if block is None and name is None and val is None:
            continue
        out.append(
            {
                "block": "" if block is None else str(block).strip(),
                "name": "" if name is None else str(name).strip(),
                "name_norm": norm(name),
                "value": safe_display_value(val),
            }
        )
    return out


# ----------------------------
# Specs
# ----------------------------
ACTIVATION_TYPES = [
    "Тип активации: ДО МЕРОПРИЯТИЯ (медиа продвижение и PR)",
    "Тип активации: ПРОДВИЖЕНИЕ НА МЕРОПРИЯТИИ",
    "Тип активации: ДОПОЛНИТЕЛЬНЫЕ АКТИВАЦИИ ПОСЛЕ МЕРОПРИЯТИЯ",
]

AUTO_FILTER_LABELS = {
    norm("ЦА (унифицированная аудитория для всех медиа, тыс. 16+)"),
    norm("Кол-во посетителей всего, тыс."),
    norm("Общий бюджет"),
    norm("Стоимость привлеченного клиента"),
    norm("Стоимость за контакт"),
    norm("Стоимость за охваченного пользователя"),
    norm("Стоимость за посетителя мероприятия"),
}

# Editable in draft only
EDITABLE_FILTER_LABELS = [
    "количество дней на фестивале/площадке",
    "Общий период размещения",
    "План посетителей (в тыс. человек)",
    "Средняя цена билета",
    "GMV",
    "Агентская комиссия",
    "Количество проданных билетов через виджит/витрину",
    "Интеграционный платеж (организатору)",
    "Продакшен фото-зоны и лайтбокса, букингсекретных артистов",
]

EDITABLE_TABLE_COLS_DRAFT = [
    "Описание",
    "Форматы",
    "Период размещения, дни",
    "Доля брендирования",
    "OTS 16+  ('000)",
    "Охват 16+* ('000)",
    "Взаимодействие с контентом",
    "Характеристики инвентаря и аудитории",
    "Хронометраж",
]


# Auto columns: robust detection (line breaks / spacing)
def is_auto_table_col(col_name: str) -> bool:
    c = norm(col_name)
    if ("ots 16+" in c) and ("с учетом доли брендирования" in c):
        return True
    if "время взаимодействия с креативом" in c:
        return True
    if ("охват 16+" in c) and ("с учетом доли брендирования" in c):
        return True
    return False


# ----------------------------
# Styling
# ----------------------------
def inject_css():
    st.markdown(
        """
<style>

/* =========================
   MTC Live Brand UI
   Palette: RED #FF0032, BLACK #000000, WHITE #FFFFFF
   ========================= */

/* App background + text */
html, body, .stApp {
  background: #000000 !important;
  color: #FFFFFF !important;
}

/* Streamlit header / toolbar (no white strip) */
header[data-testid="stHeader"],
div[data-testid="stToolbar"],
div[data-testid="stDecoration"],
div[data-testid="stStatusWidget"] {
  background: #000000 !important;
  color: #FFFFFF !important;
}

/* Global padding (wide layout kept) */
.block-container {
  /*
    Streamlit renders a fixed top toolbar/header.
    If padding-top is too small, the first row (our title) can be clipped under it,
    especially on different browser zoom levels.
  */
  /* Reduce empty space above content, but keep a small safe-area for browser zoom */
  padding-top: 2.4rem !important;
  padding-bottom: 2rem;
  padding-left: 2rem;
  padding-right: 2rem;
}

/* -------------------------
   Compact header (title + logo)
   ------------------------- */
.mtc-app-header{
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap: 1rem;
  /* Slight top padding helps prevent emoji/title clipping at some zoom levels */
  padding-top: 0.35rem;
  margin: 0 0 0.15rem 0;
}
.mtc-app-title{
  /* Slightly smaller title to keep header compact */
  font-size: 2.3rem;
  font-weight: 800;
  line-height: 1.2;
  margin: 0;
  /* Keep a small top padding (emoji is wrapped in a span for baseline control) */
  padding: 0.55rem 0 0.1rem 0;
  overflow: visible !important;
}

.mtc-app-title-icon{
  display: inline-block;
  line-height: 1.2;
  position: relative;
  top: 4px;
  margin-right: 12px;
}

/* Remove extra vertical gap that Streamlit adds around the header markdown block */
div[data-testid="stMarkdown"]:has(.mtc-app-header){
  margin-bottom: 0 !important;
  padding-bottom: 0 !important;
}
div[data-testid="stMarkdown"]:has(.mtc-app-header) > div{
  margin-bottom: 0 !important;
}
/*
  Streamlit renders `st.markdown()` and `st.image()` as separate sibling blocks,
  so a wrapper div from markdown cannot reliably style the image via descendant selectors.
  To keep the header compact and avoid the huge vertical gap, we normalize st.image globally:
  - right-align the image block
  - constrain image max-height (prevents the header row from becoming tall)
*/
/* Header logo alignment (scoped via the marker div) */
.mtc-header-right{
  width:100%;
  height:0;
  margin:0;
  padding:0;
}
.mtc-header-right + div[data-testid="stImage"]{
  display:flex;
  justify-content:flex-end;
  align-items:center;
  overflow: visible !important;
  line-height:0;
  margin:0 !important;
  padding:0 !important;
}
.mtc-header-right + div[data-testid="stImage"] img{
  display:block;
  max-height: 90px !important;
  width: auto !important;
  height: auto !important;
  max-width: 100% !important;
  object-fit: contain !important;
  border-radius: 0 !important;
}
/* Remove any blue focus/outline */
*:focus { outline: none !important; box-shadow: none !important; }

/* -------------------------
   Labels (readable, white)
   ------------------------- */
label, .stMarkdown, .stTextInput label, .stSelectbox label, .stNumberInput label {
  color: #FFFFFF !important;
  opacity: 1 !important;
  font-size: 0.9rem !important;
}

/* -------------------------
   Read-only blocks (ro_field)
   ------------------------- */
.ro-field { margin: 0 0 0.9rem 0; }
.ro-label {
  font-size: 0.9rem;
  color: #FFFFFF;
  opacity: 1;
  margin-bottom: 0.35rem;
}
.ro-value {
  border: none;
  border-radius: 0.45rem;
  padding: 10px 12px;
  line-height: 1.35rem;
  background: #FFFFFF;
  color: #000000;
}
.ro-value.auto {
  background: #FF0032;
  color: #FFFFFF;
}

/* -------------------------
   User inputs (white, no dirty border)
   BaseWeb components
   ------------------------- */

/* Input container */
div[data-baseweb="input"] > div {
  background: #FFFFFF !important;
  border: none !important;
  box-shadow: none !important;
  border-radius: 0.45rem !important;
}

/* Actual input */
div[data-baseweb="input"] input {
  background: #FFFFFF !important;
  color: #000000 !important;
  -webkit-text-fill-color: #000000 !important;
  padding: 10px 12px !important;
  border: none !important;
  box-shadow: none !important;
}

/* Select container */
div[data-baseweb="select"] > div {
  background: #FFFFFF !important;
  border: none !important;
  box-shadow: none !important;
  border-radius: 0.45rem !important;
}
div[data-baseweb="select"] span,
div[data-baseweb="select"] div {
  color: #000000 !important;
}

/* Focus state: red outline (no blue) */
div[data-baseweb="input"]:focus-within > div,
div[data-baseweb="select"]:focus-within > div {
  outline: 2px solid #FF0032 !important;
  outline-offset: 0px !important;
}

/* Disabled / AUTO fields: red, white text, not faded */
div[data-baseweb="input"] input:disabled {
  background: #FF0032 !important;
  -webkit-text-fill-color: #FFFFFF !important;
  color: #FFFFFF !important;
  opacity: 1 !important;
  padding: 10px 12px !important;
  border: none !important;
}

/* -------------------------
   Buttons (CTA)
   ------------------------- */
button[kind="primary"], button[kind="secondary"] {
  background: #FF0032 !important;
  color: #FFFFFF !important;
  border: none !important;
  box-shadow: none !important;
}
button[kind="primary"]:hover,
button[kind="secondary"]:hover,
button[kind="primary"]:active,
button[kind="secondary"]:active {
  background: #FFFFFF !important;
  color: #000000 !important;
}
button:focus { outline: 2px solid #FF0032 !important; }

/* -------------------------
   AgGrid (white table, no blue)
   ------------------------- */
.ag-theme-alpine {
  --ag-background-color: #FFFFFF;
  --ag-foreground-color: #000000;
  --ag-header-background-color: #FFFFFF;
  --ag-header-foreground-color: #000000;
  --ag-border-color: rgba(0,0,0,0.10);
  --ag-row-hover-color: rgba(255,0,50,0.08);
  --ag-selected-row-background-color: rgba(255,0,50,0.12);
  --ag-range-selection-border-color: #FF0032;
  --ag-alpine-active-color: #FF0032;
}

.ag-root-wrapper, .ag-root, .ag-body-viewport, .ag-center-cols-viewport {
  background: #FFFFFF !important;
}

/* Header bottom red line */
.ag-header {
  border-bottom: 1px solid #FF0032 !important;
}

/* Remove blue focus ring in cells */
.ag-cell-focus, .ag-cell:focus, .ag-cell:focus-within {
  outline: none !important;
  box-shadow: none !important;
}

/* Selection / range without blue */
.ag-row.ag-row-selected {
  background: rgba(255,0,50,0.12) !important;
}
.ag-cell-range-selected, .ag-cell-range-selected-1, .ag-cell-range-selected-2, .ag-cell-range-selected-3, .ag-cell-range-selected-4 {
  background: rgba(255,0,50,0.10) !important;
}


/* Cell vertical align TOP + wrap long text */
.ag-cell {
  display: flex !important;
  align-items: flex-start !important;
}
.ag-cell-value, .ag-cell-wrapper {
  white-space: normal !important;
  line-height: 1.2rem !important;
  padding-top: 6px !important;
  padding-bottom: 6px !important;
}


/* Cell vertical align TOP + wrap long text (AgGrid) */
.ag-cell {
  align-items: flex-start !important;
}
.ag-cell-wrapper {
  align-items: flex-start !important;
}
.ag-cell-value, .ag-cell-wrapper, .ag-cell-wrapper > * {
  white-space: normal !important;
  line-height: 1.2rem !important;
}
.ag-cell-value {
  padding-top: 6px !important;
  padding-bottom: 6px !important;
}


/* Legend above Calculate button */
.mtc-legend {
  display: flex;
  flex-wrap: wrap;
  gap: 14px;
  align-items: center;
  margin: 0 0 10px 0;
  color: #FFFFFF;
  font-size: 0.95rem;
}
.mtc-legend-item {
  display: inline-flex;
  gap: 8px;
  align-items: center;
}
.mtc-swatch {
  width: 14px;
  height: 14px;
  border-radius: 2px;
  display: inline-block;
}
.mtc-swatch.red {
  background: #FF0032;
  border: 1px solid #FF0032;
}
.mtc-swatch.white {
  background: #FFFFFF;
  border: 1px solid rgba(255,255,255,0.75);
}


/* Logo: prevent any clipping/cropping at fractional zoom levels */
div[data-testid="stImage"] { overflow: visible !important; }
div[data-testid="stImage"] img {
  object-fit: contain !important;
  height: auto !important;
  max-width: 100% !important;
}


/* Images: prevent clipping */
div[data-testid="stImage"] { overflow: visible !important; }




/* Ensure header row containers never clip children at fractional zoom */
div[data-testid="stHorizontalBlock"], div[data-testid="stColumn"] {
  overflow: visible !important;
}

/* -------------------------
   Header logo (no crop, right aligned)
   ------------------------- */

/* Колонка, чтобы логотип всегда был справа */
.mtc-logo-col{
  width:100%;
  display:flex;
  justify-content:flex-end;
}

/* Обёртка логотипа: даём высоту строке и выравниваем по центру */
.mtc-header-right .stImage{ margin: 0 !important; }
.mtc-logo-col .stImage{ display:flex; justify-content:flex-end; }

.mtc-logo-wrap{
  width:100%;
  display:flex;
  justify-content:flex-end;
  align-items:center;          /* ключевое */
  min-height: 0 !important; !important;            /* compact header height */
  padding-top: 0 !important;    /* убираем "подпихивание" вверх */
  overflow:visible !important;
  line-height:0;
}


/* Внутренний контейнер: ограничиваем ширину */
.mtc-logo-inner{
  display:flex;
  justify-content:flex-end;
  max-width:420px;
  width:100%;
  overflow:visible !important;
}

/* Само изображение: не растягивать по ширине, сохранять пропорции */
.mtc-logo-wrap img{
  max-height: 80px !important;
  width: auto !important;
  height:auto !important;
  max-width:100% !important;
  object-fit:contain !important;
  display:block;
  overflow:visible !important;
  border-radius: 0 !important; /* keep sharp corners */
}

/* Force Streamlit image wrapper to stick to the right edge */
.mtc-logo-inner{
  width:auto !important;
}
.mtc-logo-inner .stImage{
  margin-left:auto !important;
  margin-right:0 !important;
  display:flex !important;
  justify-content:flex-end !important;
}
.mtc-logo-inner .stImage img{
  margin-left:auto !important;
  margin-right:0 !important;
}


/* Header right column wrapper: keep logo pinned to the right edge */
.mtc-header-right{
  width:100%;
  display:flex;
  justify-content:flex-end;
  align-items:flex-start;
}


div[data-testid="stMarkdownContainer"] > h1{
  margin: 0 0 0.25rem 0 !important;
  padding: 0 !important;
  line-height: 1.15 !important;
  overflow: visible !important;
}

h1, h2, h3 { margin-bottom: 0.2rem !important; }
div[data-testid="stMarkdownContainer"]{
  overflow: visible !important;
}

/* Logo: avoid any rounded-corner clipping (handled above in .mtc-logo-wrap img) */
div[data-testid="stImage"] img{ border-radius:0 !important; }


.mtc-title-wrap{
  min-height: 80px;
  display:flex;
  align-items:center;
  padding-top: 0.25rem !important;
  overflow: visible !important;
}
</style>
        """,
        unsafe_allow_html=True,
    )
def ro_field(label: str, value, auto: bool = False):
    v = "" if value is None else value
    v_str = str(v) if v != "" else ""
    cls = "ro-value auto" if auto else "ro-value"
    st.markdown(
        f"""
<div class="ro-field">
  <div class="ro-label">{label}</div>
  <div class="{cls}">{v_str if v_str else "&nbsp;"}</div>
</div>
        """,
        unsafe_allow_html=True,
    )


# ----------------------------
# State
# ----------------------------
def ensure_state():
    st.session_state.setdefault("calculated", False)
    st.session_state.setdefault("geo", "Москва")
    st.session_state.setdefault("venue_type", "Площадка")
    st.session_state.setdefault("filter_inputs", {})
    st.session_state.setdefault("table_inputs", {})
    st.session_state.setdefault("truth_filters_rows", [])
    st.session_state.setdefault("truth_tables", {})


# ----------------------------
# Table helpers
# ----------------------------
def build_empty_table(headers_b_to_m: List[str], n_rows: int) -> pd.DataFrame:
    return pd.DataFrame([["" for _ in headers_b_to_m] for _ in range(n_rows)], columns=headers_b_to_m)


def split_tables_from_truth(df_a_to_m: pd.DataFrame, headers: List[str]) -> Dict[str, pd.DataFrame]:
    col_a = headers[0]
    cols_b_to_m = headers[1:13]  # B..M
    out: Dict[str, pd.DataFrame] = {}
    for act in ACTIVATION_TYPES:
        part = df_a_to_m[df_a_to_m[col_a].astype(str).apply(lambda x: str(x).strip()) == act]
        out[act] = part[cols_b_to_m].copy().reset_index(drop=True)
    return out


def aggrid_table(
    df: pd.DataFrame,
    editable: bool,
    dropdown_options: Dict[str, List[str]],
    height: int,
    key: str,
) -> pd.DataFrame:
    gb = GridOptionsBuilder.from_dataframe(df)

    # Make columns readable by default:
    # - flex: distribute width to show all columns without needing manual resize immediately
    # - wrap header text & auto header height: show full header labels
    gb.configure_default_column(
        resizable=True,
        sortable=False,
        filter=False,
        wrapHeaderText=True,
        autoHeaderHeight=True,
        wrapText=True,
        autoHeight=True,
        flex=1,
        minWidth=140,
    )

    editable_norms = {norm(x) for x in EDITABLE_TABLE_COLS_DRAFT}
    dropdown_norms = {norm("Описание"), norm("Форматы")}

    for col in df.columns:
        ncol = norm(col)
        auto_col = is_auto_table_col(col)

        can_edit = False
        if editable and (ncol in editable_norms) and (not auto_col):
            can_edit = True

        if ncol in dropdown_norms:
            if can_edit:
                gb.configure_column(
                    col,
                    editable=True,
                    cellEditor="agSelectCellEditor",
                    cellEditorParams={"values": dropdown_options.get(col, [])},
                )
            else:
                gb.configure_column(col, editable=False)
        else:
            gb.configure_column(col, editable=can_edit)

        if 'auto_unique_id' in ncol:
            gb.configure_column(col, hide=True)
        elif auto_col:
            gb.configure_column(
                col,
                cellStyle={"backgroundColor": "#FF0032", "color": "#FFFFFF"},
                editable=False,
            )

    gb.configure_grid_options(domLayout="normal")
    grid_options = gb.build()

    resp = AgGrid(
        df,
        gridOptions=grid_options,
        data_return_mode=DataReturnMode.FILTERED_AND_SORTED,
        update_mode=GridUpdateMode.MODEL_CHANGED,
        allow_unsafe_jscode=False,
        theme="alpine",
        height=height,
        fit_columns_on_grid_load=True,
        key=key,
    )
    return resp["data"]


# ----------------------------
# Main
# ----------------------------
def main():
    st.set_page_config(page_title="Калькулятор оценки мероприятий", layout="wide")
    inject_css()
    ensure_state()

    root_xlsx = "Калькулятор.xlsx"
    root_moscow_xlsx = "Калькулятор_Москва.xlsx"

    if not os.path.exists(root_xlsx) or not os.path.exists(root_moscow_xlsx):
        st.error(
            "Не найдены файлы в корне репозитория. "
            "Положите рядом с app_v3.py файлы: 'Калькулятор.xlsx' и 'Калькулятор_Москва.xlsx'."
        )
        return

    xlsx_mtime = file_mtime(root_xlsx)

    # Draft-time allowed sources (refresh on Excel change via cache key mtime)
    lists_options = read_lists_options_cached(root_xlsx, "Списки", xlsx_mtime)
    geo_options = lists_options.get("ГЕО", [])
    venue_options = lists_options.get("Тип площадки", [])

    descr_options = read_single_column_list_cached(root_xlsx, "Описание", 1, xlsx_mtime)
    format_options = read_formats_list_cached(root_xlsx, "Форматы", xlsx_mtime)

    # Ensure defaults (must be done BEFORE selectboxes; do not pass index => avoids yellow Streamlit warning)
    if geo_options:
        if st.session_state["geo"] not in geo_options:
            st.session_state["geo"] = "Москва" if "Москва" in geo_options else geo_options[0]
    if venue_options:
        if st.session_state["venue_type"] not in venue_options:
            st.session_state["venue_type"] = "Площадка" if "Площадка" in venue_options else venue_options[0]

    # Draft-mode allowed: only headers A..M and column A for grouping
    media_headers_a_to_m, group_counts = find_media_headers_and_groups(root_xlsx, "Медиа факторы")
    headers_b_to_m = media_headers_a_to_m[1:13]

    # Init empty tables once
    if not st.session_state["table_inputs"]:
        for act in ACTIVATION_TYPES:
            st.session_state["table_inputs"][act] = build_empty_table(headers_b_to_m, int(group_counts.get(act, 0)))

    calculated = bool(st.session_state["calculated"])
    truth_rows = st.session_state.get("truth_filters_rows", [])
    truth_tables = st.session_state.get("truth_tables", {})
    # Header (title + logo)
    # NOTE: We render the header (title + logo) as a single HTML block.
    # This avoids layout/cropping issues that can happen when trying to "wrap"
    # Streamlit widgets (st.title/st.image) with HTML opened/closed in separate
    # st.markdown calls (Streamlit renders each widget as a separate DOM block).

    # Keep logo path discovery logic unchanged
    logo_path = None
    try:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        cwd_dir = os.getcwd()
        logo_candidates = [
            "assets/MTC_Live_logo white.png",
            "./assets/MTC_Live_logo white.png",
            "MTC_Live_logo white.png",
        ]
        # Search roots: script dir, cwd, and one level up for both (helps when app is in /app
        # but assets are in repo root).
        search_roots = [
            script_dir,
            cwd_dir,
            os.path.dirname(script_dir),
            os.path.dirname(cwd_dir),
        ]

        for candidate in logo_candidates:
            # 1) relative to known roots
            for root in search_roots:
                candidate_path = os.path.join(root, candidate)
                if os.path.exists(candidate_path):
                    logo_path = candidate_path
                    break
            if logo_path:
                break

            # 2) plain relative path as-is
            if os.path.exists(candidate):
                logo_path = candidate
                break
    except Exception:
        logo_path = None

    # --- Header (title left, logo right) ---
    # Keep it pure Streamlit (st.columns + st.image) to avoid base64-HTML rendering quirks.
    h_left, h_right = st.columns([10, 2])
    with h_left:
        st.markdown(
            '<div class="mtc-app-title"><span class="mtc-app-title-icon">📊</span>Калькулятор оценки мероприятий</div>',
            unsafe_allow_html=True,
        )
    with h_right:
        # Marker node used for stable CSS targeting of the following Streamlit image block.
        # We don't wrap st.image with HTML because Streamlit renders each element as a separate block.
        st.markdown('<div class="mtc-header-right"></div>', unsafe_allow_html=True)
        if logo_path and os.path.exists(logo_path):
            st.image(logo_path, width=320)

    # Top blocks (4 columns)
    # ----------------------------
    col1, col2, col3, col4 = st.columns(4)

    # Draft helpers
    def draft_value(label: str) -> str:
        if norm(label) in AUTO_FILTER_LABELS:
            return ""
        return st.session_state["filter_inputs"].get(norm(label), "")

    def render_block_truth(block_name: str):
        """Render ALL rows from sheet 'Фильтры' for the given block, in Excel order."""
        for row in truth_rows:
            if row.get("block", "") != block_name:
                continue
            label = row.get("name", "")
            val = row.get("value", "")
            ro_field(label, val, auto=(row.get("name_norm", "") in AUTO_FILTER_LABELS))

    with col1:
        st.subheader("⚙️ Параметры")

        if calculated:
            # Render entire block from Excel (includes ЦА and other параметрические строки)
            render_block_truth("ПАРАМЕТРЫ")
            # GEO & venue type per spec: keep chosen values, read-only, NOT blue
            ro_field("ГЕО", st.session_state["geo"], auto=False)
            ro_field("Тип площадки", st.session_state["venue_type"], auto=False)
        else:
            # Draft view: ЦА is AUTO => blue disabled input (empty)
            st.text_input(
                "ЦА (унифицированная аудитория для всех медиа, тыс. 16+)",
                value="",
                disabled=True,
                key="auto_ca",
            )

            st.selectbox("ГЕО", options=geo_options if geo_options else ["Москва"], key="geo")
            st.selectbox("Тип площадки", options=venue_options if venue_options else ["Площадка"], key="venue_type")

            # Draft editable inputs (must persist)
            for lab, key in [
                ("количество дней на фестивале/площадке", "w_days"),
                ("Общий период размещения", "w_period"),
                ("План посетителей (в тыс. человек)", "w_plan_visitors"),
            ]:
                st.text_input(lab, value=str(draft_value(lab)), key=key)
                st.session_state["filter_inputs"][norm(lab)] = st.session_state.get(key, "")

    with col2:
        st.subheader("🎯 Плановый результат")
        if calculated:
            render_block_truth("ПЛАНОВЫЙ РЕЗУЛЬТАТ МЕРОПРИЯТИЯ")
        else:
            # Draft: show required fields; AUTO is blue disabled and empty
            st.text_input("Кол-во посетителей всего, тыс.", value="", disabled=True, key="auto_visitors_total")
            for lab, key in [
                ("Средняя цена билета", "w_ticket"),
                ("GMV", "w_gmv"),
                ("Агентская комиссия", "w_fee"),
                ("Количество проданных билетов через виджит/витрину", "w_widget"),
            ]:
                st.text_input(lab, value=str(draft_value(lab)), key=key)
                st.session_state["filter_inputs"][norm(lab)] = st.session_state.get(key, "")

    with col3:
        st.subheader("💰 Бюджет")
        if calculated:
            render_block_truth("БЮДЖЕТ")
        else:
            for lab, key in [
                ("Интеграционный платеж (организатору)", "w_integration"),
                ("Продакшен фото-зоны и лайтбокса, букингсекретных артистов", "w_production"),
            ]:
                st.text_input(lab, value=str(draft_value(lab)), key=key)
                st.session_state["filter_inputs"][norm(lab)] = st.session_state.get(key, "")
            st.text_input("Общий бюджет", value="", disabled=True, key="auto_total_budget")

    with col4:
        st.subheader("📈 Эффективность")
        if calculated:
            render_block_truth("ЭФФЕКТИВНОСТЬ")
        else:
            for lab, key in [
                ("Стоимость привлеченного клиента", "auto_cac"),
                ("Стоимость за контакт", "auto_cpc"),
                ("Стоимость за охваченного пользователя", "auto_cpu"),
                ("Стоимость за посетителя мероприятия", "auto_cpv"),
            ]:
                st.text_input(lab, value="", disabled=True, key=key)

    st.divider()

    # ----------------------------
    # Tables (3)
    # ----------------------------
    dropdown_options = {"Описание": descr_options, "Форматы": format_options}

    for act in ACTIVATION_TYPES:
        st.markdown(f"### {act}")

        if calculated:
            df_show = truth_tables.get(act, build_empty_table(headers_b_to_m, 0))
            _ = aggrid_table(
                df=df_show,
                editable=False,
                dropdown_options=dropdown_options,
                height=240 if len(df_show) <= 6 else 360,
                key=f"grid_truth_{norm(act)}",
            )
        else:
            df_draft = st.session_state["table_inputs"].get(act, build_empty_table(headers_b_to_m, 0))
            df_new = aggrid_table(
                df=df_draft,
                editable=True,
                dropdown_options=dropdown_options,
                height=240 if len(df_draft) <= 6 else 360,
                key=f"grid_draft_{norm(act)}",
            )
            st.session_state["table_inputs"][act] = df_new

    st.divider()

    # ----------------------------
    # Buttons
    # ----------------------------
    st.markdown(
        """<div class="mtc-legend">
  <span class="mtc-legend-item"><span class="mtc-swatch red"></span>🔴 Авторасчёт</span>
  <span class="mtc-legend-item"><span class="mtc-swatch white"></span>⬜ Ввод пользователя</span>
</div>""",
        unsafe_allow_html=True,
    )

    if st.button("🧮 Рассчитать", disabled=calculated):
        st.session_state["truth_filters_rows"] = read_filters_rows_truth(root_xlsx, "Фильтры")
        headers_truth, df_truth_a_to_m = read_media_factors_truth(root_xlsx, "Медиа факторы")
        st.session_state["truth_tables"] = split_tables_from_truth(df_truth_a_to_m, headers_truth)

        st.session_state["calculated"] = True
        st.rerun()

    # Download Москва-file "as is" (no openpyxl, no modification)
    try:
        with open(root_moscow_xlsx, "rb") as f:
            st.download_button(
                label="💾 Скачать файл (.xlsx)",
                data=f,
                file_name="Калькулятор_Москва.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    except Exception:
        st.error("Не удалось открыть файл 'Калькулятор_Москва.xlsx' для скачивания. Проверьте, что файл доступен.")


if __name__ == "__main__":
    main()
