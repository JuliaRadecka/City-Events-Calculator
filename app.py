import io
import re
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ============================================================
# App config
# ============================================================

APP_TITLE = "📊 Калькулятор оценки мероприятий"

# NOTE:
# The workbook is used ONLY as a layout + reference-data container (lists/coefficients)
# and as an export template. Calculations are executed in Python.
WORKBOOK_PATH = "Калькулятор_оценки_мероприятий_по_городам_все_города.xlsm"

# Excel working blocks (TEMPLATE)
PARAM_R1, PARAM_C1, PARAM_R2, PARAM_C2 = 20, 3, 26, 5     # C20:E26
MAIN_R1, MAIN_C1, MAIN_R2, MAIN_C2 = 27, 3, 134, 15       # C27:O134
BOT_R1, BOT_C1, BOT_R2, BOT_C2 = 135, 3, 146, 18          # C135:R146 (R=18)


# Known fills (ARGB) from the template
KNOWN_GREYS = {"FFD9D9D9", "FFBFBFBF", "FFE7E6E6", "FFDDDDDD"}
KNOWN_BLUES = {"FFBDD7EE", "FF9DC3E6", "FFB4C6E7", "FFB7DEE8"}


# ============================================================
# Helpers: Excel basics
# ============================================================


def a1(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def cell_fill_argb(cell) -> str:
    try:
        f = cell.fill
        if f is None or f.patternType is None:
            return ""
        c = f.fgColor
        if c is None:
            return ""
        if c.type == "rgb" and c.rgb:
            return c.rgb
    except Exception:
        return ""
    return ""


def is_grey_input(cell) -> bool:
    argb = cell_fill_argb(cell)
    if argb in KNOWN_GREYS:
        return True
    # fallback: non-white non-blue treated as grey input in this template
    if argb and argb not in KNOWN_BLUES and argb != "FFFFFFFF":
        return True
    return False


def is_blue_auto(cell) -> bool:
    return cell_fill_argb(cell) in KNOWN_BLUES


def coerce_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


# ============================================================
# Reference lists ("Списки")
# ============================================================


def extract_city_list(ws_lists) -> List[str]:
    cities: List[str] = []
    r = 2
    while True:
        v = ws_lists[f"A{r}"].value
        if v is None or str(v).strip() == "":
            break
        cities.append(str(v).strip())
        r += 1
    return cities


def extract_type_ploshadki(ws_lists) -> List[str]:
    opts: List[str] = []
    for r in range(2, 200):
        v = ws_lists[f"W{r}"].value
        if v is None or str(v).strip() == "":
            break
        opts.append(str(v).strip())
    return opts


def extract_union_formats(ws_lists) -> List[str]:
    # union of all format columns used in the template lists
    cols = ["B", "C", "D", "E", "F", "H", "I", "K", "L", "M", "N", "O", "P", "Q"]
    out: set[str] = set()
    for col in cols:
        empty_streak = 0
        for r in range(2, 500):
            v = ws_lists[f"{col}{r}"].value
            if v is None or str(v).strip() == "":
                empty_streak += 1
                if r > 30 and empty_streak > 25:
                    break
                continue
            empty_streak = 0
            out.add(str(v).strip())
    return sorted(out)


def extract_simple_col(ws, col_letter: str, start_row: int = 2, max_rows: int = 500) -> List[str]:
    out: List[str] = []
    for r in range(start_row, max_rows + 1):
        v = ws[f"{col_letter}{r}"].value
        if v is None or str(v).strip() == "":
            continue
        out.append(str(v).strip())
    return out


# ============================================================
# Excel-like formula evaluation engine (minimal, for this template)
# ============================================================


@dataclass
class EvalContext:
    wb: Any
    sheet: str
    values: Dict[str, Any]  # A1 -> scalar


_CELL_RE = re.compile(r"\$?[A-Z]{1,3}\$?\d+")


def _strip_dollars(ref: str) -> str:
    return ref.replace("$", "")


def _parse_range(rng: str) -> Tuple[str, str]:
    a, b = rng.split(":")
    return _strip_dollars(a), _strip_dollars(b)


def _expand_range(a: str, b: str) -> List[str]:
    col_a = re.findall(r"[A-Z]{1,3}", a)[0]
    row_a = int(re.findall(r"\d+", a)[0])
    col_b = re.findall(r"[A-Z]{1,3}", b)[0]
    row_b = int(re.findall(r"\d+", b)[0])
    ca = column_index_from_string(col_a)
    cb = column_index_from_string(col_b)
    out: List[str] = []
    for r in range(min(row_a, row_b), max(row_a, row_b) + 1):
        for c in range(min(ca, cb), max(ca, cb) + 1):
            out.append(a1(r, c))
    return out


def _get(ctx: EvalContext, ref: str) -> Any:
    ref = _strip_dollars(ref)
    return ctx.values.get(ref)


def _to_number(v: Any) -> Optional[float]:
    return coerce_float(v)


def _excel_sum(args: List[Any]) -> float:
    total = 0.0
    for v in args:
        if isinstance(v, list):
            total += _excel_sum(v)
        else:
            num = _to_number(v)
            if num is not None:
                total += num
    return total


def _excel_max(args: List[Any]) -> float:
    nums: List[float] = []
    for v in args:
        if isinstance(v, list):
            for x in v:
                n = _to_number(x)
                if n is not None:
                    nums.append(n)
        else:
            n = _to_number(v)
            if n is not None:
                nums.append(n)
    return max(nums) if nums else 0.0


def _excel_if(cond: Any, a: Any, b: Any) -> Any:
    return a if bool(cond) else b


def _excel_iferror(expr: Any, fallback: Any) -> Any:
    return fallback if isinstance(expr, Exception) else expr


def _excel_countif(values: List[Any], criterion: Any) -> int:
    crit = str(criterion)
    # Very small subset: exact match only (used in this template for flags)
    return sum(1 for v in values if str(v) == crit)


def _excel_vlookup(ctx: EvalContext, lookup_value: Any, table_range: str, col_index: int, exact: bool = True) -> Any:
    # table_range may include a sheet prefix like: "'[1]Коэфы качество контакта'!$A$2:$G$22"
    # We accept: <sheet>!<range> and ignore [1] marker.
    sheet = ctx.sheet
    rng = table_range
    if "!" in table_range:
        sh, rng = table_range.split("!", 1)
        sh = sh.strip().strip("'")
        sh = sh.replace("[1]", "")
        sheet = sh

    a_ref, b_ref = _parse_range(rng)
    ws = ctx.wb[sheet]

    # Build rows
    a_col = re.findall(r"[A-Z]{1,3}", a_ref)[0]
    a_row = int(re.findall(r"\d+", a_ref)[0])
    b_col = re.findall(r"[A-Z]{1,3}", b_ref)[0]
    b_row = int(re.findall(r"\d+", b_ref)[0])
    ca = column_index_from_string(a_col)
    cb = column_index_from_string(b_col)

    lookup_str = str(lookup_value).strip() if lookup_value is not None else ""

    for r in range(a_row, b_row + 1):
        key = ws.cell(r, ca).value
        if str(key).strip() == lookup_str:
            target_c = ca + (col_index - 1)
            if target_c < ca or target_c > cb:
                return None
            return ws.cell(r, target_c).value
    return None


def _transform_formula(formula: str) -> str:
    """Transform an Excel formula string to a restricted Python expression.

    Supported functions: SUM, IF, IFERROR, MAX, COUNTIF, VLOOKUP
    Supported operators: + - * / ^, comparisons.
    """
    f = formula.strip()
    if f.startswith("="):
        f = f[1:]

    # power
    f = f.replace("^", "**")

    # TRUE/FALSE
    f = re.sub(r"\bTRUE\b", "True", f, flags=re.IGNORECASE)
    f = re.sub(r"\bFALSE\b", "False", f, flags=re.IGNORECASE)

    # Replace sheet refs marker [1]
    f = f.replace("[1]", "")

    # Ranges: A1:B2 -> rng('A1','B2')
    f = re.sub(
        r"(\$?[A-Z]{1,3}\$?\d+)\s*:\s*(\$?[A-Z]{1,3}\$?\d+)",
        lambda m: f"RNG('{_strip_dollars(m.group(1))}','{_strip_dollars(m.group(2))}')",
        f,
    )

    # Cell refs: $A$1 -> CELL('A1') (but avoid those already inside quotes)
    def repl_cell(m):
        ref = _strip_dollars(m.group(0))
        return f"CELL('{ref}')"

    # Temporarily protect quoted strings
    strings: List[str] = []

    def protect_str(m):
        strings.append(m.group(0))
        return f"__STR{len(strings)-1}__"

    f = re.sub(r'"[^"]*"', protect_str, f)
    f = _CELL_RE.sub(repl_cell, f)

    # restore strings
    for i, s in enumerate(strings):
        f = f.replace(f"__STR{i}__", s)

    # Map functions
    f = re.sub(r"\bSUM\s*\(", "SUM(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bMAX\s*\(", "MAX(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bIFERROR\s*\(", "IFERROR(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bIF\s*\(", "IF(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bCOUNTIF\s*\(", "COUNTIF(", f, flags=re.IGNORECASE)
    f = re.sub(r"\bVLOOKUP\s*\(", "VLOOKUP(", f, flags=re.IGNORECASE)

    return f


def eval_formula(ctx: EvalContext, formula: str) -> Any:
    expr = _transform_formula(formula)

    def CELL(ref: str) -> Any:
        return _get(ctx, ref)

    def RNG(a_ref: str, b_ref: str) -> List[Any]:
        return [_get(ctx, r) for r in _expand_range(a_ref, b_ref)]

    def SUM(*args):
        return _excel_sum(list(args))

    def MAX(*args):
        return _excel_max(list(args))

    def IF(cond, a, b):
        return _excel_if(cond, a, b)

    def IFERROR(x, fallback):
        # In our evaluation we raise exceptions only for parsing/runtime issues.
        return x if not isinstance(x, Exception) else fallback

    def COUNTIF(values, criterion):
        if not isinstance(values, list):
            values = [values]
        return _excel_countif(values, criterion)

    def VLOOKUP(lookup_value, table_range, col_index, exact=True):
        return _excel_vlookup(ctx, lookup_value, table_range, int(col_index), bool(exact))

    env = {
        "CELL": CELL,
        "RNG": RNG,
        "SUM": SUM,
        "MAX": MAX,
        "IF": IF,
        "IFERROR": IFERROR,
        "COUNTIF": COUNTIF,
        "VLOOKUP": VLOOKUP,
    }

    try:
        # Restricted eval (no builtins)
        return eval(expr, {"__builtins__": {}}, env)
    except Exception as e:
        return e


def recalc_block(ctx: EvalContext, ws, addr_list: List[str], max_iters: int = 20) -> None:
    """Iteratively evaluate formulas for a set of addresses until stable."""
    formulas: Dict[str, str] = {}
    for addr in addr_list:
        v = ws[addr].value
        if isinstance(v, str) and v.startswith("="):
            formulas[addr] = v

    for _ in range(max_iters):
        changed = 0
        for addr, f in formulas.items():
            val_new = eval_formula(ctx, f)
            # Excel-like: IFERROR can return ""; keep as-is
            val_old = ctx.values.get(addr)
            if isinstance(val_new, Exception):
                val_new = None
            if val_new != val_old:
                ctx.values[addr] = val_new
                changed += 1
        if changed == 0:
            break


# ============================================================
# UI: build blocks
# ============================================================


def read_block_values(ws, r1: int, c1: int, r2: int, c2: int) -> pd.DataFrame:
    rows: List[List[Any]] = []
    for r in range(r1, r2 + 1):
        row: List[Any] = []
        for c in range(c1, c2 + 1):
            row.append(ws.cell(r, c).value)
        rows.append(row)
    cols = [get_column_letter(c) for c in range(c1, c2 + 1)]
    idx = list(range(r1, r2 + 1))
    return pd.DataFrame(rows, columns=cols, index=idx)


def read_block_style(ws_style, r1: int, c1: int, r2: int, c2: int) -> Tuple[pd.DataFrame, pd.DataFrame]:
    grey: List[List[bool]] = []
    blue: List[List[bool]] = []
    for r in range(r1, r2 + 1):
        row_g: List[bool] = []
        row_b: List[bool] = []
        for c in range(c1, c2 + 1):
            cell = ws_style.cell(r, c)
            row_g.append(is_grey_input(cell))
            row_b.append(is_blue_auto(cell))
        grey.append(row_g)
        blue.append(row_b)
    cols = [get_column_letter(c) for c in range(c1, c2 + 1)]
    idx = list(range(r1, r2 + 1))
    return pd.DataFrame(grey, columns=cols, index=idx), pd.DataFrame(blue, columns=cols, index=idx)


def style_df(df: pd.DataFrame, grey_mask: pd.DataFrame, blue_mask: pd.DataFrame):
    def _apply(_):
        out = pd.DataFrame("", index=df.index, columns=df.columns)
        out[grey_mask] = "background-color: #D9D9D9;"  # light grey
        out[blue_mask] = "background-color: #9DC3E6;"  # Excel-like light blue
        return out

    return df.style.apply(_apply, axis=None)


def inject_css():
    st.markdown(
        """
<style>
  div[data-testid="stVerticalBlock"] div[data-testid="stTextInput"] input,
  div[data-testid="stVerticalBlock"] div[data-testid="stNumberInput"] input,
  div[data-testid="stVerticalBlock"] div[data-testid="stSelectbox"] div {
    max-width: 250px;
  }
  .block-title {
    font-weight: 700;
    font-size: 1.1rem;
    margin-bottom: 0.35rem;
  }
  .calc-area {
    padding: 0.5rem 0.75rem;
    border: 1px solid rgba(49, 51, 63, 0.2);
    border-radius: 0.5rem;
  }
</style>
""",
        unsafe_allow_html=True,
    )


# ============================================================
# App
# ============================================================


def main():
    st.set_page_config(page_title=APP_TITLE, layout="wide")
    inject_css()
    st.title(APP_TITLE)

    wb = load_workbook(WORKBOOK_PATH, data_only=False, keep_vba=True)
    if "TEMPLATE" not in wb.sheetnames:
        st.error("Не найден лист TEMPLATE в книге.")
        return
    if "Списки" not in wb.sheetnames:
        st.error("Не найден лист 'Списки' в книге.")
        return

    ws = wb["TEMPLATE"]
    ws_lists = wb["Списки"]

    # reference lists
    cities = extract_city_list(ws_lists)
    type_pl = extract_type_ploshadki(ws_lists)
    union_formats = extract_union_formats(ws_lists)

    # Optional channel lists
    tv_channels = extract_simple_col(wb["ТВ Каналы"], "A") if "ТВ Каналы" in wb.sheetnames else []
    radio_stations = extract_simple_col(wb["Радиостанции"], "A") if "Радиостанции" in wb.sheetnames else []

    # Build initial value map from the template sheet (values + inputs)
    if "calc_values" not in st.session_state:
        vals: Dict[str, Any] = {}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                addr = cell.coordinate
                v = cell.value
                # keep formulas as-is; computed values will be stored in ctx.values
                if isinstance(v, str) and v.startswith("="):
                    vals[addr] = None
                else:
                    vals[addr] = v
        st.session_state["calc_values"] = vals

    # --------------------------------------------------------
    # Top area: Parameters + Results blocks (side-by-side)
    # --------------------------------------------------------

    left, mid1, mid2, right = st.columns([1, 1, 1, 1], gap="large")

    # Parameters (⚙️)
    with left:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">⚙️Параметры</div>', unsafe_allow_html=True)

        # 3.1: CA (E21) auto (blue)
        # We'll compute it during recalculation; for now show current value.
        ca_val = st.session_state["calc_values"].get("E21")
        st.text_input(
            "ЦА (унифицированная аудитория для всех медиа, тыс. 16+)",
            value="" if ca_val is None else str(ca_val),
            disabled=True,
        )

        geo_val = st.selectbox("ГЕО", options=cities if cities else [""], index=0)
        p1, p2 = st.columns([1, 1])
        with p1:
            type_pl_val = st.selectbox("Тип площадки", options=type_pl if type_pl else [""], index=0)
        with p2:
            days_val = st.number_input("Кол-во дней на фестивале/площадке", value=0.0, step=1.0)
        period_val = st.number_input("Общий период размещения", value=0.0, step=1.0)
        visitors_val = st.number_input("План посетителей (в тыс. человек)", value=0.0, step=1.0)

        st.markdown('</div>', unsafe_allow_html=True)

    # Planned results
    with mid1:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">ПЛАНОВЫЙ РЕЗУЛЬТАТ МЕРОПРИЯТИЯ</div>', unsafe_allow_html=True)
        st.text_input("Кол-во посетителей всего, тыс.", value=str(visitors_val), disabled=True)
        ticket_price = st.number_input("Средняя цена билета", value=0.0, step=1.0)
        gmv = st.number_input("GMV", value=0.0, step=1.0)
        fee = st.number_input("Агентская комиссия", value=0.0, step=0.1)
        widget_tickets = st.number_input("Количество проданных билетов через виджит/витрину", value=0.0, step=1.0)
        st.markdown('</div>', unsafe_allow_html=True)

    # Budget
    with mid2:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">💰 БЮДЖЕТ</div>', unsafe_allow_html=True)
        integ_fee = st.number_input("Интеграционный платеж (организатору)", value=0.0, step=1.0)
        production = st.number_input(
            "Продакшен фото-зоны и лайтбокса, букинг секретных артистов",
            value=0.0,
            step=1.0,
        )
        total_budget = integ_fee + production
        st.text_input("Общий бюджет", value=f"{total_budget}", disabled=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Efficiency
    with right:
        st.markdown('<div class="calc-area">', unsafe_allow_html=True)
        st.markdown('<div class="block-title">📈ЭФФЕКТИВНОСТЬ</div>', unsafe_allow_html=True)
        # These are calculated in Python after media totals are computed.
        cpa = st.session_state.get("eff_cpa")
        cpc = st.session_state.get("eff_cpc")
        cpr = st.session_state.get("eff_cpr")
        cpv = st.session_state.get("eff_cpv")
        st.text_input("Стоимость привлеченного клиента", value="" if cpa is None else str(cpa), disabled=True)
        st.text_input("Стоимость за контакт", value="" if cpc is None else str(cpc), disabled=True)
        st.text_input("Стоимость за охваченного пользователя", value="" if cpr is None else str(cpr), disabled=True)
        st.text_input("Стоимость за посетителя мероприятия", value="" if cpv is None else str(cpv), disabled=True)
        st.markdown('</div>', unsafe_allow_html=True)

    st.divider()

    # --------------------------------------------------------
    # Media factors table
    # --------------------------------------------------------

    st.subheader("Медиа факторы")

    # Header row from row 27 (C:O). Data from row 28.
    header_df = read_block_values(ws, MAIN_R1, MAIN_C1, MAIN_R1, MAIN_C2)
    headers = [str(v) if v is not None else get_column_letter(c) for v, c in zip(header_df.iloc[0].tolist(), range(MAIN_C1, MAIN_C2 + 1))]

    data_df = read_block_values(ws, MAIN_R1 + 1, MAIN_C1, MAIN_R2, MAIN_C2)
    data_df.columns = headers

    # Style masks from template, shifted for data rows
    ws_style = wb["TEMPLATE"]
    grey_mask, blue_mask = read_block_style(ws_style, MAIN_R1 + 1, MAIN_C1, MAIN_R2, MAIN_C2)
    grey_mask.columns = headers
    blue_mask.columns = headers

    # Editable only on grey cells; data_editor supports per-column disabling only.
    # We will accept edits but enforce writeback by mask.
    display_df = data_df.fillna("")

    edited_df = st.data_editor(
        display_df,
        use_container_width=True,
        height=720,
        column_config={
            # Formats column: attempt to detect the header that contains "Формат"
            **({h: st.column_config.SelectboxColumn(h, options=union_formats, required=False) for h in headers if "ФОРМ" in str(h).upper()}),
        },
        key="media_editor",
    )

    # ТВ/Радио selectors (4.4)
    with st.expander("ТВ и Радио: выбор каналов / радиостанций", expanded=False):
        st.multiselect("ТВ: каналы", options=tv_channels, key="tv_channels")
        st.multiselect("Радио: радиостанции", options=radio_stations, key="radio_stations")

    # --------------------------------------------------------
    # Calculate
    # --------------------------------------------------------

    calc_clicked = st.button("🧮 Рассчитать", type="primary")

    # Always visible download button (generated after calc; if not calculated yet, exports inputs only)
    download_placeholder = st.empty()

    if calc_clicked:
        # Write user inputs back into ctx.values at the exact Excel addresses the template uses.
        vals = st.session_state["calc_values"].copy()

        # Parameters mapping (template): E22:E26, E21 formula.
        vals["E22"] = geo_val
        vals["E23"] = type_pl_val
        vals["E24"] = int(days_val) if float(days_val).is_integer() else float(days_val)
        vals["E25"] = int(period_val) if float(period_val).is_integer() else float(period_val)
        vals["E26"] = float(visitors_val)

        # Planned result / budget inputs are outside the declared blocks in the user message,
        # but we keep them as internal-only right now.
        st.session_state["plan_ticket_price"] = ticket_price
        st.session_state["plan_gmv"] = gmv
        st.session_state["plan_fee"] = fee
        st.session_state["plan_widget_tickets"] = widget_tickets
        st.session_state["budget_integ"] = integ_fee
        st.session_state["budget_prod"] = production

        # Main table edits: map edited values back by row/col to TEMPLATE addresses.
        # MAIN data rows are 28..134, columns C..O.
        for i, excel_row in enumerate(range(MAIN_R1 + 1, MAIN_R2 + 1)):
            for j, excel_col in enumerate(range(MAIN_C1, MAIN_C2 + 1)):
                addr = a1(excel_row, excel_col)
                # enforce only grey inputs
                if not grey_mask.iloc[i, j]:
                    continue
                col_name = headers[j]
                v = edited_df.iloc[i, j]
                v = None if v == "" else v
                vals[addr] = v

        ctx = EvalContext(wb=wb, sheet="TEMPLATE", values=vals)

        # Recalculate the necessary set: E21 + MAIN blue cells + bottom block formulas.
        calc_cells: List[str] = ["E21"]

        # MAIN block (C27:O134) formulas are in the template; we recalc all formula cells there.
        for r in range(MAIN_R1, MAIN_R2 + 1):
            for c in range(MAIN_C1, MAIN_C2 + 1):
                addr = a1(r, c)
                if isinstance(ws[addr].value, str) and ws[addr].value.startswith("="):
                    calc_cells.append(addr)

        # Bottom block
        for r in range(BOT_R1, BOT_R2 + 1):
            for c in range(BOT_C1, BOT_C2 + 1):
                addr = a1(r, c)
                if isinstance(ws[addr].value, str) and ws[addr].value.startswith("="):
                    calc_cells.append(addr)

        recalc_block(ctx, ws, calc_cells)

        st.session_state["calc_values"] = ctx.values

        # Update CA display
        st.session_state["ca_value"] = ctx.values.get("E21")

        # Simple efficiency metrics computed from totals (fallback):
        # - Cost per visitor = total budget / visitors
        visitors = coerce_float(visitors_val) or 0.0
        st.session_state["eff_cpv"] = (total_budget / visitors) if visitors else None

        # If we can locate totals in the bottom block, use them to calculate CPM-like metrics.
        # We try to detect total OTS and total Reach in bottom block by header labels.
        try:
            bot_vals = ctx.values
            # Heuristic: sum of O column (O = 15) in MAIN corresponds to Reach (?)
            # Template-specific; keep safe.
            pass
        except Exception:
            pass

        st.success("Расчёт выполнен.")

    # --------------------------------------------------------
    # Bottom block display (read from calc_values after recalc)
    # --------------------------------------------------------

    st.subheader("Итоги / нижний блок")
    vals_map = st.session_state["calc_values"]
    bot_display_rows: List[List[Any]] = []
    for r in range(BOT_R1, BOT_R2 + 1):
        row: List[Any] = []
        for c in range(BOT_C1, BOT_C2 + 1):
            addr = a1(r, c)
            v = vals_map.get(addr)
            row.append("" if v is None else v)
        bot_display_rows.append(row)
    bot_cols = [get_column_letter(c) for c in range(BOT_C1, BOT_C2 + 1)]
    bot_df = pd.DataFrame(bot_display_rows, columns=bot_cols, index=list(range(BOT_R1, BOT_R2 + 1)))
    bot_grey, bot_blue = read_block_style(ws_style, BOT_R1, BOT_C1, BOT_R2, BOT_C2)
    st.dataframe(style_df(bot_df, bot_grey, bot_blue), use_container_width=True, height=360)

    # --------------------------------------------------------
    # Export (.xlsx)
    # --------------------------------------------------------

    def build_export_xlsx() -> bytes:
        wb_out = load_workbook(WORKBOOK_PATH, data_only=False, keep_vba=False)
        ws_out = wb_out["TEMPLATE"]

        # Write all computed values back (values only; no formulas) for user-visible cells.
        for addr, v in st.session_state["calc_values"].items():
            if addr not in ws_out:
                continue
            if isinstance(ws_out[addr].value, str) and ws_out[addr].value.startswith("="):
                ws_out[addr].value = v
            else:
                # For inputs we also overwrite to keep consistency
                if addr in {"E22", "E23", "E24", "E25", "E26"}:
                    ws_out[addr].value = v

        out = io.BytesIO()
        wb_out.save(out)
        out.seek(0)
        return out.getvalue()

    xlsx_bytes = build_export_xlsx()
    download_placeholder.download_button(
        label="💾 Скачать файл (.xlsx)",
        data=xlsx_bytes,
        file_name="Калькулятор_оценки_мероприятий.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    main()
